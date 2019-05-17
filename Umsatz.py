from openpyxl import Workbook
from openpyxl import load_workbook
import os

cwd = os.getcwd()
print(cwd)

directory = 'C:\\Users\\alper\\Desktop\\Praktikum\\04_Workshop'
#wb = Workbook()
folder = os.listdir(directory)

y = ' '
liste = []
c_sieben = 0
c_acht = 0
c_neun = 0
c_zehn = 0
zaehler_dateien = 0
sheet_count = 0

#Change directory
#os.chdir(".\Rechnung\")

#laeuft durch alle Excel Dateien
for files in folder: 
    if files.endswith(".xlsx"):        
        
        #wb = load_workbook(filename)
        #wb2 = load_workbook('Rechnung_EricIdle.xlsx')
        wb2 = load_workbook(os.path.join(directory, files))
        #print(wb2.sheetnames)
        sheet = wb2.active
        
        #Zaehle Anzahl der Sheets
        for anzahl_sheet in wb2:
            sheet_count +=1
            print("Sheet_Count", sheet_count)
            print(wb2.sheetnames)
        
        
#       #laeuft durch Sheets
        for sheets in wb2:
            zaehler_dateien += 1
            
            #Name
            b3 = sheet['B3'].value
            b4 = sheet['B4'].value
            liste = [b3, b4]
            
            #Werte
            c7 = sheet['C7'].value
            c8 = sheet['C8'].value
            c9 = sheet['C9'].value
            c10 = sheet['C10'].value
            
            c_sieben += c7
            c_acht += c8
            c_neun += c9
            c_zehn += c10
            #c10 = sheet.cell(10,3)
            
            liste.append(b3)
            liste.append(b4)
            
            print("Zaehler", zaehler_dateien)
            print(c7)
            print(c8)
            print(c9)
            print(c10)
            print("File gefunden")
            #continue
            print(c_sieben)
            print(c_acht)
            print(c_neun)
            print(c_zehn)
            

                        
                        
                        #neues excel-File erzeugen
    else:
        print("Datei konnte nicht gefunden werden")
        print("\n")
        continue

#        print("Erstelle neues workbook")
wb3 = Workbook()  
ws = wb3.active
ws.title =str("Umsatz")

            #Kunden Namen einlesen
ws['A1'] = 'Es wurden xy Dateien eingelesen'
ws['A3'] = 'Artikel'
ws['A4'] = 'Briefumschlag'
ws['A5'] = 'Bleistift'
ws['A6'] = 'Lineael'
ws['A7'] = 'Textmarker'

ws['B3'] = 'Gesamtzahl'
ws['B4'] = c_sieben
ws['B5'] = c_acht
ws['B6'] = c_neun
ws['B7'] = c_zehn

wb3.save('C:\\Users\\alper\\Desktop\\Praktikum\\04_Workshop'+'Umsatz.xlsx')

print("Es wurden", zaehler_dateien, "Dateien eingelesen")
print("\n")
print("Artikel", "        ", "Gesamtzahl")
print("Briefumschlag", "        ", c_sieben)
print("Bleistift", "        ", c_acht)
print("Lineal", "        ", c_neun)
print("Textmartker", "        ", c_zehn)
print("\n")
kundenliste = input(str("Kundenliste ausgeben y/n ? "))
if kundenliste == 'y':
    print(b3, b4)
    print(liste)
